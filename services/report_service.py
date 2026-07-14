"""
Report Service — Excel and PDF generation using openpyxl and reportlab.
"""
import io
from datetime import date
from typing import Optional, List


# ---------------------------------------------------------------------------
# Excel Reports
# ---------------------------------------------------------------------------
def generate_students_excel(students) -> io.BytesIO:
    """Generate Excel report for a list of students."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Students'

    # Header style
    header_font = Font(bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill(start_color='1e3a5f', end_color='1e3a5f', fill_type='solid')
    center = Alignment(horizontal='center', vertical='center')
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    # Title
    ws.merge_cells('A1:K1')
    ws['A1'] = 'EduTrack Pro — Student Report'
    ws['A1'].font = Font(bold=True, size=14, color='1e3a5f')
    ws['A1'].alignment = center
    ws.row_dimensions[1].height = 30

    ws.merge_cells('A2:K2')
    ws['A2'] = f'Generated on: {date.today().strftime("%d %B %Y")}'
    ws['A2'].alignment = center
    ws.row_dimensions[2].height = 20

    # Column headers
    headers = ['#', 'Student ID', 'Full Name', 'Roll Number', 'Department',
               'Year', 'Semester', 'Email', 'Phone', 'Attendance %', 'CGPA']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = thin_border

    ws.row_dimensions[4].height = 20

    # Data rows
    alt_fill = PatternFill(start_color='EEF2FF', end_color='EEF2FF', fill_type='solid')
    for idx, student in enumerate(students, 1):
        row = idx + 4
        data = [
            idx,
            student.student_id,
            student.full_name,
            student.roll_number,
            student.department.name if student.department else '',
            student.year,
            student.semester,
            student.email,
            student.phone or '',
            f'{student.attendance_percentage():.1f}%',
            student.calculate_cgpa(),
        ]
        fill = alt_fill if idx % 2 == 0 else None
        for col, value in enumerate(data, 1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center' if col in (1, 2, 4, 6, 7, 10, 11) else 'left')
            if fill:
                cell.fill = fill

    # Column widths
    widths = [5, 15, 25, 15, 30, 8, 10, 30, 15, 15, 10]
    for col, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = width

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def generate_attendance_excel(dept_id: Optional[int], subject_id: Optional[int],
                               month_str: str) -> io.BytesIO:
    """Generate attendance report Excel."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from models import Student, Attendance, Subject, Department
    from sqlalchemy import func
    from extensions import db

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Attendance Report'

    try:
        year, month = map(int, month_str.split('-'))
    except (ValueError, AttributeError):
        year, month = date.today().year, date.today().month

    ws['A1'] = 'EduTrack Pro — Attendance Report'
    ws['A1'].font = Font(bold=True, size=14, color='1e3a5f')
    ws['A2'] = f'Month: {date(year, month, 1).strftime("%B %Y")} | Generated: {date.today().strftime("%d %B %Y")}'

    headers = ['#', 'Student ID', 'Student Name', 'Roll No', 'Present', 'Absent', 'Late', 'Total', 'Percentage', 'Status']
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='2d6a4f', end_color='2d6a4f', fill_type='solid')

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')

    query = Student.query.filter_by(is_active=True)
    if dept_id:
        query = query.filter_by(department_id=dept_id)
    students = query.order_by(Student.roll_number).all()

    for idx, student in enumerate(students, 1):
        att_query = Attendance.query.filter_by(student_id=student.id)
        if subject_id:
            att_query = att_query.filter_by(subject_id=subject_id)
        att_query = att_query.filter(
            func.extract('year', Attendance.date) == year,
            func.extract('month', Attendance.date) == month,
        )
        records = att_query.all()
        present = sum(1 for r in records if r.status == 'present')
        absent = sum(1 for r in records if r.status == 'absent')
        late = sum(1 for r in records if r.status == 'late')
        total = len(records)
        pct = round(((present + late) / total) * 100, 1) if total > 0 else 0
        status = 'OK' if pct >= 75 else 'LOW'

        row_data = [idx, student.student_id, student.full_name, student.roll_number,
                    present, absent, late, total, f'{pct}%', status]
        for col, val in enumerate(row_data, 1):
            ws.cell(row=idx + 4, column=col, value=val)

    ws.column_dimensions['C'].width = 25
    ws.column_dimensions['D'].width = 15
    for col in ['A', 'B', 'E', 'F', 'G', 'H', 'I', 'J']:
        ws.column_dimensions[col].width = 12

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def generate_marks_excel(dept_id: Optional[int], semester: Optional[int],
                          academic_year: str) -> io.BytesIO:
    """Generate marks report Excel."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    from models import Student, Mark

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Marks Report'

    ws['A1'] = f'EduTrack Pro — Marks Report | AY: {academic_year}'
    ws['A1'].font = Font(bold=True, size=14)

    headers = ['#', 'Student ID', 'Name', 'Roll No', 'Department', 'Subject',
               'Exam Type', 'Obtained', 'Max', 'Percentage', 'Grade', 'Pass/Fail']
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='6d28d9', end_color='6d28d9', fill_type='solid')

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')

    query = Mark.query
    if semester:
        query = query.filter_by(semester=semester, academic_year=academic_year)
    if dept_id:
        from models import Subject
        subj_ids = [s.id for s in Subject.query.filter_by(department_id=dept_id).all()]
        if subj_ids:
            query = query.filter(Mark.subject_id.in_(subj_ids))

    marks = query.order_by(Mark.student_id).all()

    for idx, m in enumerate(marks, 1):
        ws.cell(row=idx + 3, column=1, value=idx)
        ws.cell(row=idx + 3, column=2, value=m.student.student_id if m.student else '')
        ws.cell(row=idx + 3, column=3, value=m.student.full_name if m.student else '')
        ws.cell(row=idx + 3, column=4, value=m.student.roll_number if m.student else '')
        ws.cell(row=idx + 3, column=5, value=m.student.department.name if m.student and m.student.department else '')
        ws.cell(row=idx + 3, column=6, value=m.subject.name if m.subject else '')
        ws.cell(row=idx + 3, column=7, value=m.exam_type)
        ws.cell(row=idx + 3, column=8, value=m.marks_obtained)
        ws.cell(row=idx + 3, column=9, value=m.max_marks)
        ws.cell(row=idx + 3, column=10, value=f'{m.percentage:.1f}%')
        ws.cell(row=idx + 3, column=11, value=m.grade)
        ws.cell(row=idx + 3, column=12, value='Pass' if m.is_pass else 'Fail')

    for col in range(1, 13):
        ws.column_dimensions[get_column_letter(col)].width = 15
    ws.column_dimensions['C'].width = 25

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


# ---------------------------------------------------------------------------
# PDF Reports
# ---------------------------------------------------------------------------
def generate_students_pdf(students, dept=None) -> io.BytesIO:
    """Generate PDF student list report using reportlab."""
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.units import cm

    output = io.BytesIO()
    doc = SimpleDocTemplate(output, pagesize=landscape(A4),
                            leftMargin=1*cm, rightMargin=1*cm,
                            topMargin=1.5*cm, bottomMargin=1.5*cm)
    styles = getSampleStyleSheet()
    story = []

    # Title
    title_style = ParagraphStyle('Title', parent=styles['Title'],
                                  fontSize=18, textColor=colors.HexColor('#1e3a5f'))
    story.append(Paragraph('EduTrack Pro — Student Report', title_style))
    dept_name = dept.name if dept else 'All Departments'
    story.append(Paragraph(f'Department: {dept_name} | Date: {date.today().strftime("%d %B %Y")}',
                           styles['Normal']))
    story.append(Spacer(1, 0.5*cm))

    # Table data
    headers = ['#', 'Student ID', 'Name', 'Roll No', 'Department', 'Year', 'Email', 'Attendance%', 'CGPA']
    data = [headers]
    for idx, s in enumerate(students, 1):
        data.append([
            str(idx), s.student_id, s.full_name, s.roll_number,
            s.department.name[:20] if s.department else '', str(s.year),
            s.email[:30], f'{s.attendance_percentage():.1f}%', str(s.calculate_cgpa()),
        ])

    table = Table(data, repeatRows=1)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1e3a5f')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 9),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#EEF2FF')]),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
    ]))

    story.append(table)
    doc.build(story)
    output.seek(0)
    return output


def generate_marksheet_pdf(student, semester: Optional[int], academic_year: str) -> io.BytesIO:
    """Generate individual student marksheet PDF."""
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, HRFlowable
    from reportlab.lib.units import cm
    from models import Mark

    output = io.BytesIO()
    doc = SimpleDocTemplate(output, pagesize=A4,
                            leftMargin=2*cm, rightMargin=2*cm,
                            topMargin=2*cm, bottomMargin=2*cm)
    styles = getSampleStyleSheet()
    story = []

    blue = colors.HexColor('#1e3a5f')
    accent = colors.HexColor('#4f46e5')

    # Header
    title_style = ParagraphStyle('T', parent=styles['Title'], fontSize=20, textColor=blue)
    story.append(Paragraph('EduTrack Pro', title_style))
    story.append(Paragraph('Official Mark Sheet', ParagraphStyle('S', parent=styles['Normal'],
                           fontSize=13, textColor=accent)))
    story.append(HRFlowable(width='100%', color=blue, thickness=2))
    story.append(Spacer(1, 0.3*cm))

    # Student info table
    info = [
        ['Student ID:', student.student_id, 'Name:', student.full_name],
        ['Roll No:', student.roll_number, 'Department:', student.department.name if student.department else ''],
        ['Year:', str(student.year), 'Semester:', str(semester or student.semester)],
        ['Academic Year:', academic_year, 'CGPA:', str(student.calculate_cgpa())],
    ]
    info_table = Table(info, colWidths=[4*cm, 5*cm, 4*cm, 5*cm])
    info_table.setStyle(TableStyle([
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ('FONTNAME', (2, 0), (2, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('ROWBACKGROUNDS', (0, 0), (-1, -1), [colors.HexColor('#F8FAFF'), colors.white]),
        ('GRID', (0, 0), (-1, -1), 0.3, colors.lightgrey),
    ]))
    story.append(info_table)
    story.append(Spacer(1, 0.5*cm))

    # Marks table
    marks_query = Mark.query.filter_by(student_id=student.id, academic_year=academic_year)
    if semester:
        marks_query = marks_query.filter_by(semester=semester)
    marks = marks_query.order_by(Mark.subject_id).all()

    headers = ['Subject', 'Exam Type', 'Obtained', 'Max', '%', 'Grade', 'Status']
    data = [headers]
    for m in marks:
        data.append([
            m.subject.name if m.subject else '',
            m.exam_type.title(),
            str(m.marks_obtained),
            str(m.max_marks),
            f'{m.percentage:.1f}%',
            m.grade,
            'Pass' if m.is_pass else 'Fail',
        ])

    marks_table = Table(data, repeatRows=1)
    marks_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), blue),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('ALIGN', (0, 1), (0, -1), 'LEFT'),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#EEF2FF')]),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
    ]))

    story.append(Paragraph('Academic Performance', ParagraphStyle('H2', parent=styles['Heading2'],
                           textColor=blue)))
    story.append(marks_table)
    story.append(Spacer(1, 1*cm))
    story.append(Paragraph(
        f'Generated on {date.today().strftime("%d %B %Y")} | EduTrack Pro Student Management System',
        ParagraphStyle('Footer', parent=styles['Normal'], fontSize=8, textColor=colors.grey)
    ))

    doc.build(story)
    output.seek(0)
    return output
