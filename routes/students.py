"""
Student Management routes — Full CRUD with search, filter, pagination.
"""
import os
from flask import (Blueprint, render_template, redirect, url_for, flash,
                   request, current_app, send_from_directory)
from flask_login import login_required, current_user
from werkzeug.utils import secure_filename
from sqlalchemy import or_

from extensions import db
from models import Student, Department, User
from services.student_service import (
    generate_student_id, get_paginated_students, search_students
)
from services import cloudinary_service

students_bp = Blueprint('students', __name__, url_prefix='/students')


def _allowed_image(filename: str) -> bool:
    return ('.' in filename and
            filename.rsplit('.', 1)[1].lower() in
            current_app.config['ALLOWED_IMAGE_EXTENSIONS'])


@students_bp.route('/')
@login_required
def list_students():
    page = request.args.get('page', 1, type=int)
    per_page = current_app.config.get('STUDENTS_PER_PAGE', 15)
    search = request.args.get('search', '').strip()
    dept_id = request.args.get('dept', type=int)
    year = request.args.get('year', type=int)
    sort = request.args.get('sort', 'name')

    query = Student.query.filter_by(is_active=True)

    # Search
    if search:
        query = query.filter(
            or_(
                Student.full_name.ilike(f'%{search}%'),
                Student.roll_number.ilike(f'%{search}%'),
                Student.student_id.ilike(f'%{search}%'),
                Student.email.ilike(f'%{search}%'),
            )
        )

    # Filters
    if dept_id:
        query = query.filter_by(department_id=dept_id)
    if year:
        query = query.filter_by(year=year)

    # Sorting
    if sort == 'name':
        query = query.order_by(Student.full_name)
    elif sort == 'roll':
        query = query.order_by(Student.roll_number)
    elif sort == 'newest':
        query = query.order_by(Student.created_at.desc())
    elif sort == 'dept':
        query = query.join(Department).order_by(Department.name)

    pagination = query.paginate(page=page, per_page=per_page, error_out=False)
    departments = Department.query.order_by(Department.name).all()

    return render_template(
        'students/list.html',
        students=pagination.items,
        pagination=pagination,
        departments=departments,
        search=search,
        selected_dept=dept_id,
        selected_year=year,
        sort=sort,
    )


@students_bp.route('/add', methods=['GET', 'POST'])
@login_required
def add_student():
    if not (current_user.is_admin() or current_user.is_teacher()):
        flash('Permission denied.', 'danger')
        return redirect(url_for('students.list_students'))

    departments = Department.query.order_by(Department.name).all()

    if request.method == 'POST':
        # Validate required fields
        full_name = request.form.get('full_name', '').strip()
        roll_number = request.form.get('roll_number', '').strip()
        department_id = request.form.get('department_id', type=int)
        year = request.form.get('year', 1, type=int)
        semester = request.form.get('semester', 1, type=int)
        email = request.form.get('email', '').strip()
        phone = request.form.get('phone', '').strip()
        address = request.form.get('address', '').strip()
        dob_str = request.form.get('date_of_birth', '')

        # Validation
        errors = []
        if not full_name:
            errors.append('Full name is required.')
        if not roll_number:
            errors.append('Roll number is required.')
        if not department_id:
            errors.append('Department is required.')
        if not email:
            errors.append('Email is required.')
        if Student.query.filter_by(roll_number=roll_number).first():
            errors.append(f'Roll number "{roll_number}" already exists.')
        if Student.query.filter_by(email=email).first():
            errors.append(f'Email "{email}" already exists.')

        if errors:
            for e in errors:
                flash(e, 'danger')
            return render_template('students/add.html', departments=departments)

        # Handle photo upload
        photo_filename = None
        photo_url = None
        photo_public_id = None
        if 'photo' in request.files:
            photo = request.files['photo']
            if photo and photo.filename and _allowed_image(photo.filename):
                if current_app.config.get('CLOUDINARY_CLOUD_NAME'):
                    # Upload to Cloudinary
                    try:
                        photo_url, photo_public_id = cloudinary_service.upload_file(
                            photo, folder='edutrack/photos', resource_type='image'
                        )
                    except Exception as e:
                        flash(f'Photo upload failed: {e}', 'warning')
                else:
                    # Fallback: local storage (dev without Cloudinary creds)
                    filename = secure_filename(photo.filename)
                    student_id_str = generate_student_id()
                    photo_filename = f"{student_id_str}_{filename}"
                    photo_path = os.path.join(current_app.config['UPLOAD_FOLDER'], 'photos', photo_filename)
                    photo.save(photo_path)

        # Parse DOB
        from datetime import date as date_type
        dob = None
        if dob_str:
            try:
                dob = date_type.fromisoformat(dob_str)
            except ValueError:
                pass

        student = Student(
            student_id=generate_student_id(),
            full_name=full_name,
            roll_number=roll_number,
            department_id=department_id,
            year=year,
            semester=semester,
            email=email,
            phone=phone,
            address=address,
            date_of_birth=dob,
            photo_filename=photo_filename,
            photo_url=photo_url,
            photo_public_id=photo_public_id,
        )
        db.session.add(student)
        db.session.commit()

        flash(f'Student "{full_name}" added successfully! ID: {student.student_id}', 'success')
        return redirect(url_for('students.student_detail', student_id=student.id))

    return render_template('students/add.html', departments=departments)


@students_bp.route('/<int:student_id>')
@login_required
def student_detail(student_id):
    student = db.session.get(Student, student_id)
    if not student:
        flash('Student not found.', 'danger')
        return redirect(url_for('students.list_students'))

    # Attendance summary per subject
    from models import Attendance, Subject
    subjects = Subject.query.filter_by(department_id=student.department_id).all()
    attendance_summary = []
    for subj in subjects:
        total = Attendance.query.filter_by(student_id=student.id, subject_id=subj.id).count()
        present = Attendance.query.filter_by(
            student_id=student.id, subject_id=subj.id
        ).filter(Attendance.status.in_(['present', 'late'])).count()
        pct = round((present / total) * 100, 1) if total > 0 else 0
        attendance_summary.append({
            'subject': subj,
            'total': total,
            'present': present,
            'percentage': pct,
            'low': pct < 75 and total > 0,
        })

    # Marks summary
    from models import Mark
    marks = Mark.query.filter_by(student_id=student.id).order_by(
        Mark.semester, Mark.subject_id
    ).all()

    # Uploaded files
    from models import UploadedFile
    files = UploadedFile.query.filter_by(student_id=student.id).order_by(
        UploadedFile.uploaded_at.desc()
    ).all()

    return render_template(
        'students/detail.html',
        student=student,
        attendance_summary=attendance_summary,
        marks=marks,
        files=files,
        overall_attendance=student.attendance_percentage(),
        cgpa=student.calculate_cgpa(),
    )


@students_bp.route('/<int:student_id>/edit', methods=['GET', 'POST'])
@login_required
def edit_student(student_id):
    if not (current_user.is_admin() or current_user.is_teacher()):
        flash('Permission denied.', 'danger')
        return redirect(url_for('students.list_students'))

    student = db.session.get(Student, student_id)
    if not student:
        flash('Student not found.', 'danger')
        return redirect(url_for('students.list_students'))

    departments = Department.query.order_by(Department.name).all()

    if request.method == 'POST':
        student.full_name = request.form.get('full_name', student.full_name).strip()
        student.roll_number = request.form.get('roll_number', student.roll_number).strip()
        student.department_id = request.form.get('department_id', student.department_id, type=int)
        student.year = request.form.get('year', student.year, type=int)
        student.semester = request.form.get('semester', student.semester, type=int)
        student.email = request.form.get('email', student.email).strip()
        student.phone = request.form.get('phone', '').strip()
        student.address = request.form.get('address', '').strip()

        dob_str = request.form.get('date_of_birth', '')
        if dob_str:
            from datetime import date as date_type
            try:
                student.date_of_birth = date_type.fromisoformat(dob_str)
            except ValueError:
                pass

        # Handle photo update
        if 'photo' in request.files:
            photo = request.files['photo']
            if photo and photo.filename and _allowed_image(photo.filename):
                if current_app.config.get('CLOUDINARY_CLOUD_NAME'):
                    # Delete old Cloudinary photo if exists
                    if student.photo_public_id:
                        cloudinary_service.delete_file(student.photo_public_id, resource_type='image')
                    try:
                        photo_url, photo_public_id = cloudinary_service.upload_file(
                            photo, folder='edutrack/photos', resource_type='image'
                        )
                        student.photo_url = photo_url
                        student.photo_public_id = photo_public_id
                        student.photo_filename = None  # clear legacy local filename
                    except Exception as e:
                        flash(f'Photo upload failed: {e}', 'warning')
                else:
                    # Fallback: local storage
                    filename = secure_filename(photo.filename)
                    photo_filename = f"{student.student_id}_{filename}"
                    photo_path = os.path.join(current_app.config['UPLOAD_FOLDER'], 'photos', photo_filename)
                    photo.save(photo_path)
                    student.photo_filename = photo_filename

        db.session.commit()
        flash(f'Student "{student.full_name}" updated successfully.', 'success')
        return redirect(url_for('students.student_detail', student_id=student.id))

    return render_template('students/edit.html', student=student, departments=departments)


@students_bp.route('/<int:student_id>/delete', methods=['POST'])
@login_required
def delete_student(student_id):
    if not current_user.is_admin():
        flash('Admin access required.', 'danger')
        return redirect(url_for('students.list_students'))

    student = db.session.get(Student, student_id)
    if not student:
        flash('Student not found.', 'danger')
        return redirect(url_for('students.list_students'))

    name = student.full_name
    # Soft delete
    student.is_active = False
    db.session.commit()
    flash(f'Student "{name}" has been removed.', 'success')
    return redirect(url_for('students.list_students'))


@students_bp.route('/photos/<filename>')
def student_photo(filename):
    """Serve locally stored photos (legacy fallback for dev without Cloudinary)."""
    upload_folder = current_app.config['UPLOAD_FOLDER']
    return send_from_directory(os.path.join(upload_folder, 'photos'), filename)


@students_bp.route('/import', methods=['POST'])
@login_required
def import_students():
    if not (current_user.is_admin() or current_user.is_teacher()):
        flash('Permission denied.', 'danger')
        return redirect(url_for('students.list_students'))

    file = request.files.get('file')
    clear_existing = bool(request.form.get('clear_existing'))

    if not file or not file.filename:
        flash('No file selected.', 'danger')
        return redirect(url_for('students.list_students'))

    if not (file.filename.endswith('.xlsx') or file.filename.endswith('.xls')):
        flash('Invalid file format. Please upload an Excel file (.xlsx or .xls).', 'danger')
        return redirect(url_for('students.list_students'))

    try:
        import openpyxl
        from datetime import datetime, date as date_type
        
        # Load workbook
        wb = openpyxl.load_workbook(file, data_only=True)
        sheet = wb.active

        # Find header row
        header_row = None
        for r in range(1, 10):
            row_vals = [str(sheet.cell(row=r, column=c).value or '').strip().lower() for c in range(1, 15)]
            if 'roll number' in row_vals or 'fullname' in [rv.replace(' ', '') for rv in row_vals] or 'rollno' in [rv.replace(' ', '') for rv in row_vals]:
                header_row = r
                break
                
        if not header_row:
            header_row = 1

        # Map headers
        cols = {}
        for c in range(1, 20):
            val = str(sheet.cell(row=header_row, column=c).value or '').strip().lower().replace(' ', '').replace('_', '')
            if not val:
                continue
            if 'roll' in val:
                cols['roll_number'] = c
            elif 'name' in val:
                cols['full_name'] = c
            elif 'email' in val:
                cols['email'] = c
            elif 'dept' in val or 'department' in val:
                cols['department'] = c
            elif 'year' in val:
                cols['year'] = c
            elif 'sem' in val:
                cols['semester'] = c
            elif 'phone' in val:
                cols['phone'] = c
            elif 'address' in val:
                cols['address'] = c
            elif 'dob' in val or 'birth' in val:
                cols['dob'] = c

        # Verify minimum headers are present
        if 'roll_number' not in cols or 'full_name' not in cols:
            flash('Failed to parse Excel. Ensure "Roll Number" and "Full Name" headers exist.', 'danger')
            return redirect(url_for('students.list_students'))

        # If clear_existing is checked, delete existing students and their linked User profile
        if clear_existing:
            # Delete students (which cascade deletes attendance, marks, files due to relationship settings)
            Student.query.delete()
            # Delete users that are role='student'
            User.query.filter_by(role='student').delete()
            db.session.commit()

        imported_count = 0
        updated_count = 0

        for r in range(header_row + 1, sheet.max_row + 1):
            # Read unique roll number & full name
            roll_val = sheet.cell(row=r, column=cols['roll_number']).value
            name_val = sheet.cell(row=r, column=cols['full_name']).value
            if not roll_val or not name_val:
                continue

            roll_number = str(roll_val).strip()
            full_name = str(name_val).strip()
            
            # Read or generate email
            email = ''
            if 'email' in cols:
                email = str(sheet.cell(row=r, column=cols['email']).value or '').strip()
            if not email:
                email = f"{roll_number.lower()}@university.edu"

            # Parse optional fields
            year = 1
            if 'year' in cols:
                try:
                    year = int(sheet.cell(row=r, column=cols['year']).value or 1)
                except (ValueError, TypeError):
                    pass

            semester = 1
            if 'semester' in cols:
                try:
                    semester = int(sheet.cell(row=r, column=cols['semester']).value or 1)
                except (ValueError, TypeError):
                    pass

            phone = ''
            if 'phone' in cols:
                phone = str(sheet.cell(row=r, column=cols['phone']).value or '').strip()

            address = ''
            if 'address' in cols:
                address = str(sheet.cell(row=r, column=cols['address']).value or '').strip()

            dob = None
            if 'dob' in cols:
                dob_val = sheet.cell(row=r, column=cols['dob']).value
                if isinstance(dob_val, (datetime, date_type)):
                    dob = dob_val
                elif isinstance(dob_val, str) and dob_val:
                    try:
                        dob = datetime.strptime(dob_val.strip(), '%Y-%m-%d').date()
                    except ValueError:
                        pass

            # Look up or create department
            dept_name = 'General'
            if 'department' in cols:
                dept_name = str(sheet.cell(row=r, column=cols['department']).value or 'General').strip()

            dept = Department.query.filter((Department.name.ilike(dept_name)) | (Department.code.ilike(dept_name))).first()
            if not dept:
                # Create a new department
                dept_code = dept_name[:5].upper().replace(' ', '')
                dept = Department(name=dept_name, code=dept_code)
                db.session.add(dept)
                db.session.commit()

            # Handle existing students if we are not clearing
            existing_student = None
            if not clear_existing:
                existing_student = Student.query.filter(
                    (Student.roll_number == roll_number) | (Student.email == email)
                ).first()

            if existing_student:
                # Update existing student details
                existing_student.full_name = full_name
                existing_student.department_id = dept.id
                existing_student.year = year
                existing_student.semester = semester
                existing_student.phone = phone
                existing_student.address = address
                if dob:
                    existing_student.date_of_birth = dob
                updated_count += 1
            else:
                # Create default user account for student login
                username = roll_number.lower()
                user = User.query.filter_by(username=username).first()
                if not user:
                    user = User(username=username, email=email, role='student')
                    user.set_password('Student@123')
                    db.session.add(user)
                    db.session.commit()

                # Add new student
                student = Student(
                    student_id=generate_student_id(),
                    full_name=full_name,
                    roll_number=roll_number,
                    department_id=dept.id,
                    year=year,
                    semester=semester,
                    email=email,
                    phone=phone,
                    address=address,
                    date_of_birth=dob,
                    user_id=user.id
                )
                db.session.add(student)
                imported_count += 1

        db.session.commit()
        msg = f"Excel import complete. "
        if imported_count > 0:
            msg += f"Imported {imported_count} new student(s). "
        if updated_count > 0:
            msg += f"Updated {updated_count} student(s). "
        flash(msg, 'success')

    except Exception as e:
        db.session.rollback()
        flash(f'Failed to import Excel file: {str(e)}', 'danger')

    return redirect(url_for('students.list_students'))
